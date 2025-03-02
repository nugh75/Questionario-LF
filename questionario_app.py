import os
import io
import json
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from flask_migrate import Migrate

app = Flask(__name__)
app.config['SECRET_KEY'] = 'chiave_segreta_molto_sicura_da_cambiare'
db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'valutazioni.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Definizione delle aree
AREE = {
    'MOTIVAZIONE': [1, 17, 24, 31, 33, 45],
    'USO DELLE RISORSE ACCADEMICHE': [2, 7, 12, 27, 46, 60],
    'ELABORAZIONE DELLE INFORMAZIONI': [3, 10, 18, 22, 35, 41],
    'GESTIONE DEL TEMPO': [4, 8, 11, 23, 51, 54],
    'STRATEGIE PER SVOLGERE PROVE': [5, 21, 30, 36, 43, 57],
    'CONCENTRAZIONE': [6, 13, 25, 40, 47, 58],
    'SELEZIONE DEI CONCETTI PRINCIPALI': [9, 16, 19, 44, 48, 55],
    'ATTEGGIAMENTO': [14, 29, 32, 39, 42, 59],
    'AUTO-MONITORAGGIO': [15, 20, 26, 38, 49, 52],
    'ANSIA': [28, 34, 37, 50, 53, 56]
}

AREE_INVERSE = {}
for area, numeri in AREE.items():
    for numero in numeri:
        AREE_INVERSE[numero] = area

# Definizione delle domande
MAPPA_DOMANDE = {
    1: "Anche quando i materiali di studio sono noiosi e poco interessanti, riesco a continuare a studiare finché non finisco.",
    2: "Quando trovo difficile completare un compito assegnato in un corso, non chiedo aiuto.",
    3: "Cerco di trovare relazioni tra ciò che sto imparando e ciò che già so.",
    4: "Trovo difficile rispettare una pianificazione temporale dello studio da me stabilita.",
    5: "Nello svolgere prove di verifica o esami, mi rendo conto di aver frainteso ciò che viene richiesto e ottengo voti più bassi per questo motivo.",
    6: "Quando studio, mi concentro totalmente.",
    7: "Quando ho difficoltà in uno o più corsi, mi vergogno troppo per ammetterlo con qualcuno.",
    8: "Quando decido di studiare, mi prefiggo di dedicare un tempo specifico e lo rispetto.",
    9: "Durante le lezioni, ho difficoltà a capire cosa sia abbastanza importante da scrivere negli appunti.",
    10: "Per aiutarmi a ricordare i nuovi concetti che imparo durante una lezione, cerco di trovare una loro applicazione pratica.",
    11: "Quando si tratta di studiare, tendo spesso a rimandare.",
    12: "Se ho problemi con un compito scritto, cerco aiuto usufruendo delle risorse disponibili nella mia università, come i laboratori di scrittura, i centri per l'apprendimento o i servizi di tutorato didattico.",
    13: "Trovo difficile mantenere la concentrazione mentre svolgo le attività proposte nei corsi che seguo.",
    14: "Studio solo gli argomenti che mi interessano.",
    15: "Quando mi preparo per una prova di verifica o un esame, formulo domande che penso potrebbero essermi poste.",
    16: "Ho difficoltà a individuare i concetti importanti durante la lettura dei materiali di studio.",
    17: "Quando il materiale di studio è difficile, o mi arrendo o studio solo le parti facili.",
    18: "Per aiutarmi a imparare il materiale presentato durante le lezioni, lo relaziono alle mie conoscenze generali.",
    19: "Nei materiali di studio sono presenti così tanti dettagli che mi è difficile individuare i concetti principali.",
    20: "Rileggo i miei appunti prima della lezione successiva.",
    21: "Ho difficoltà ad adattare le mie strategie di studio ai diversi tipi di insegnamenti.",
    22: "Riformulo ciò che sto studiando con parole mie.",
    23: "Rimando nel tempo lo studio più di quanto dovrei.",
    24: "Anche se ho difficoltà in un corso, riesco ad automotivarmi per completare il lavoro da svolgere.",
    25: "Quando studio, la mia mente si distrae molto.",
    26: "Mentre leggo, mi fermo ogni tanto per riesaminare o ripassare mentalmente ciò che è stato detto.",
    27: "Non mi sento a mio agio nel chiedere un aiuto ai/alle docenti dei corsi che seguo.",
    28: "Mi sento molto in ansia quando svolgo una prova di verifica o un esame importante.",
    29: "Partecipo volentieri alle lezioni.",
    30: "Quando mi preparo per una prova di verifica o un esame, ho difficoltà a capire cosa fare per imparare il materiale di studio.",
    31: "Anche se un compito non mi piace, sono capace di impegnarmi a portarlo a termine.",
    32: "Preferirei non frequentare l'università.",
    33: "Mi propongo degli obiettivi riguardo ai voti che voglio ottenere alle prove di verifica o agli esami.",
    34: "Mentre svolgo una prova di verifica o un esame, la preoccupazione di ottenere risultati insoddisfacenti interferisce con la mia concentrazione.",
    35: "Cerco di capire come ciò che sto studiando potrebbe applicarsi alla mia vita quotidiana.",
    36: "Ho difficoltà a capire esattamente che cosa mi viene chiesto nelle domande d'esame o nelle prove di verifica.",
    37: "Mi preoccupa il pensiero di non riuscire a terminare il mio percorso di studi.",
    38: "Per assicurarmi di comprendere il materiale di studio di un corso, rileggo i miei appunti prima della lezione successiva.",
    39: "Non mi interessa ottenere una buona formazione generale, voglio solo ottenere un buon lavoro.",
    40: "Trovo difficile mantenere l'attenzione durante le lezioni.",
    41: "Cerco di mettere in relazione ciò che sto studiando con le mie esperienze personali.",
    42: "Non mi piace la maggior parte delle attività proposte nelle lezioni che seguo.",
    43: "Quando svolgo prove di verifica o esami scritti, rileggo ciò che ho scritto per assicurarmi di aver esposto e supportato adeguatamente le mie idee principali.",
    44: "Quando studio, mi sembra di perdermi nei dettagli e di non cogliere le informazioni importanti.",
    45: "Non mi impegno molto nell'ottenere ottimi risultati nei corsi che seguo.",
    46: "Se un corso è troppo difficile per me, chiedo aiuto ai servizi di tutorato didattico universitario.",
    47: "Mi distraggo molto facilmente dai miei studi.",
    48: "È difficile per me decidere cosa sia importante sottolineare in un testo.",
    49: "Per verificare la mia comprensione del materiale di studio di un corso, formulo delle possibili domande d'esame e provo a rispondere.",
    50: "Anche quando sono ben preparato per una prova di verifica o un esame, provo molta ansia.",
    51: "Dedico più tempo allo studio degli argomenti che trovo più difficili.",
    52: "Mi auto-esamino per verificare se capisco ciò che sto studiando.",
    53: "I corsi di alcune materie di studio mi mettono ansia.",
    54: "Finisco per studiare tutto all'ultimo momento prima di ogni prova di verifica o esame.",
    55: "Quando ascolto le lezioni, riesco a selezionare le informazioni importanti.",
    56: "Quando studio, la preoccupazione di ottenere scarsi risultati interferisce con la mia concentrazione.",
    57: "Ottengo scarsi risultati alle prove di verifica e agli esami perché trovo difficile organizzare il lavoro mio studio in un breve periodo di tempo.",
    58: "Se mi distraggo durante una lezione, riesco a rimettere a fuoco l'attenzione.",
    59: "Secondo me, ciò che viene insegnato nei corsi che seguo ha scarso valore.",
    60: "Quando non capisco come usare una strategia presentata in uno dei corsi che seguo, chiedo aiuto a un compagno/una compagna di corso."
}

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)

class Domanda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    testo = db.Column(db.Text, nullable=False)
    area = db.Column(db.String(50), nullable=False)
    ordine = db.Column(db.Integer, nullable=False)
    etichetta = db.Column(db.String(100))
    feedback_basso = db.Column(db.Text)
    feedback_medio = db.Column(db.Text)
    feedback_alto = db.Column(db.Text)
    attiva = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class ClusterRisposte(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    area = db.Column(db.String(50), nullable=False)
    min_valore = db.Column(db.Float, nullable=False)
    max_valore = db.Column(db.Float, nullable=False)
    feedback = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Risposta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    risposte = db.Column(db.JSON, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', name='fk_risposta_user'))
    # Medie per area
    media_motivazione = db.Column(db.Float)
    media_risorse = db.Column(db.Float)
    media_elaborazione = db.Column(db.Float)
    media_tempo = db.Column(db.Float)
    media_strategie = db.Column(db.Float)
    media_concentrazione = db.Column(db.Float)
    media_selezione = db.Column(db.Float)
    media_atteggiamento = db.Column(db.Float)
    media_monitoraggio = db.Column(db.Float)
    media_ansia = db.Column(db.Float)

    def calcola_medie(self):
        r = self.risposte
        def calcola_media_sicura(area):
            try:
                valori = [float(r.get(f'q{i}', 0)) for i in AREE[area]]
                return sum(valori) / len(valori) if valori else 0.0
            except (ValueError, TypeError):
                return 0.0

        self.media_motivazione = calcola_media_sicura('MOTIVAZIONE')
        self.media_risorse = calcola_media_sicura('USO DELLE RISORSE ACCADEMICHE')
        self.media_elaborazione = calcola_media_sicura('ELABORAZIONE DELLE INFORMAZIONI')
        self.media_tempo = calcola_media_sicura('GESTIONE DEL TEMPO')
        self.media_strategie = calcola_media_sicura('STRATEGIE PER SVOLGERE PROVE')
        self.media_concentrazione = calcola_media_sicura('CONCENTRAZIONE')
        self.media_selezione = calcola_media_sicura('SELEZIONE DEI CONCETTI PRINCIPALI')
        self.media_atteggiamento = calcola_media_sicura('ATTEGGIAMENTO')
        self.media_monitoraggio = calcola_media_sicura('AUTO-MONITORAGGIO')
        self.media_ansia = calcola_media_sicura('ANSIA')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def get_feedback(media, area):
    cluster = ClusterRisposte.query.filter(
        ClusterRisposte.area == area,
        ClusterRisposte.min_valore <= media,
        ClusterRisposte.max_valore >= media
    ).first()
    
    if cluster:
        return cluster.feedback
    return 'Nessun feedback disponibile per questo punteggio'

def get_area_domanda(numero):
    """Helper per ottenere l'area di una domanda"""
    for area, numeri in AREE.items():
        if numero in numeri:
            return area
    return None

@app.route('/', methods=['GET', 'POST'])
@app.route('/questionario', methods=['GET', 'POST'])
def questionario():
    if request.method == 'POST':
        risposte = {}
        for i in range(1, 61):
            chiave = f'q{i}'
            risposte[chiave] = request.form.get(chiave)
        
        nuova_risposta = Risposta(risposte=risposte)
        nuova_risposta.calcola_medie()
        
        db.session.add(nuova_risposta)
        db.session.commit()
        
        # Ottieni i feedback specifici per ogni area
        feedback = {
            'motivazione': get_feedback(nuova_risposta.media_motivazione, 'MOTIVAZIONE'),
            'risorse': get_feedback(nuova_risposta.media_risorse, 'USO DELLE RISORSE ACCADEMICHE'),
            'elaborazione': get_feedback(nuova_risposta.media_elaborazione, 'ELABORAZIONE DELLE INFORMAZIONI'),
            'tempo': get_feedback(nuova_risposta.media_tempo, 'GESTIONE DEL TEMPO'),
            'strategie': get_feedback(nuova_risposta.media_strategie, 'STRATEGIE PER SVOLGERE PROVE'),
            'concentrazione': get_feedback(nuova_risposta.media_concentrazione, 'CONCENTRAZIONE'),
            'selezione': get_feedback(nuova_risposta.media_selezione, 'SELEZIONE DEI CONCETTI PRINCIPALI'),
            'atteggiamento': get_feedback(nuova_risposta.media_atteggiamento, 'ATTEGGIAMENTO'),
            'monitoraggio': get_feedback(nuova_risposta.media_monitoraggio, 'AUTO-MONITORAGGIO'),
            'ansia': get_feedback(nuova_risposta.media_ansia, 'ANSIA')
        }
        
        return render_template('risultati.html', 
                          risposta_id=nuova_risposta.id,
                          medie={
                              'motivazione': nuova_risposta.media_motivazione,
                              'risorse': nuova_risposta.media_risorse,
                              'elaborazione': nuova_risposta.media_elaborazione,
                              'tempo': nuova_risposta.media_tempo,
                              'strategie': nuova_risposta.media_strategie,
                              'concentrazione': nuova_risposta.media_concentrazione,
                              'selezione': nuova_risposta.media_selezione,
                              'atteggiamento': nuova_risposta.media_atteggiamento,
                              'monitoraggio': nuova_risposta.media_monitoraggio,
                              'ansia': nuova_risposta.media_ansia
                          },
                          feedback=feedback)
    
    # Prendi tutte le domande dal database, ordinate per numero
    domande = Domanda.query.filter_by(attiva=True).order_by(Domanda.ordine).all()
    return render_template('questionario.html', domande=domande)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and check_password_hash(user.password_hash, request.form['password']):
            login_user(user)
            return redirect(url_for('admin_dashboard'))
        flash('Credenziali non valide')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/admin')
@login_required
def admin_dashboard():
    risposte = Risposta.query.order_by(Risposta.timestamp.desc()).all()
    stats = get_stats(risposte)
    return render_template('admin/dashboard.html', risposte=risposte, stats=stats)

@app.route('/admin/clusters', methods=['GET', 'POST'])
@login_required
def gestione_clusters():
    if request.method == 'POST':
        azione = request.form.get('azione')
        
        # Validazione dei valori min e max
        if azione in ['aggiungi', 'modifica']:
            try:
                min_valore = float(request.form.get('min_valore'))
                max_valore = float(request.form.get('max_valore'))
                area = request.form.get('area')
                feedback = request.form.get('feedback')
                
                if not (1 <= min_valore <= 5 and 1 <= max_valore <= 5):
                    flash('I valori devono essere compresi tra 1 e 5', 'error')
                    return redirect(url_for('gestione_clusters'))
                
                if min_valore >= max_valore:
                    flash('Il valore minimo deve essere minore del valore massimo', 'error')
                    return redirect(url_for('gestione_clusters'))
                
                # Verifica sovrapposizioni con altri cluster della stessa area
                existing_clusters = ClusterRisposte.query.filter_by(area=area).all()
                cluster_id = request.form.get('cluster_id', type=int)
                
                for cluster in existing_clusters:
                    # Salta il controllo per il cluster che stiamo modificando
                    if cluster_id and cluster.id == cluster_id:
                        continue
                    
                    if (min_valore <= cluster.max_valore and max_valore >= cluster.min_valore):
                        flash('Il range si sovrappone con un altro cluster esistente', 'error')
                        return redirect(url_for('gestione_clusters'))
                
            except ValueError:
                flash('Valori non validi per min_valore o max_valore', 'error')
                return redirect(url_for('gestione_clusters'))
        
        if azione == 'aggiungi':
            cluster = ClusterRisposte(
                area=area,
                min_valore=min_valore,
                max_valore=max_valore,
                feedback=feedback
            )
            db.session.add(cluster)
            flash('Cluster aggiunto con successo', 'success')
            
        elif azione == 'modifica':
            cluster_id = request.form.get('cluster_id', type=int)
            cluster = ClusterRisposte.query.get_or_404(cluster_id)
            
            cluster.area = area
            cluster.min_valore = min_valore
            cluster.max_valore = max_valore
            cluster.feedback = feedback
            flash('Cluster modificato con successo', 'success')
            
        elif azione == 'elimina':
            cluster_id = request.form.get('cluster_id', type=int)
            cluster = ClusterRisposte.query.get_or_404(cluster_id)
            db.session.delete(cluster)
            flash('Cluster eliminato con successo', 'success')
        
        db.session.commit()
        return redirect(url_for('gestione_clusters'))
    
    # Recupera tutti i cluster ordinati per area e valore minimo
    clusters = ClusterRisposte.query.order_by(ClusterRisposte.area, ClusterRisposte.min_valore).all()
    return render_template('admin/gestione_clusters.html', 
                         clusters=clusters, 
                         aree=AREE.keys())

@app.route('/admin/export_excel')
@login_required
def download_excel():
    output = None
    try:
        output = export_to_excel()
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='risposte_questionario.xlsx',
        )
    except Exception as e:
        app.logger.error(f"Errore durante l'esportazione Excel: {str(e)}")
        flash("Si è verificato un errore durante l'esportazione")
        if output:
            output.close()
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/tabella_risposte', methods=['GET', 'POST'])
@login_required
def tabella_risposte():
    global MAPPA_DOMANDE, AREE, AREE_INVERSE
    
    # Process POST request for question management
    if request.method == 'POST':
        azione = request.form.get('azione')
        
        if azione == 'aggiungi_domanda':
            testo = request.form.get('testo', '').strip()
            area = request.form.get('area', '').strip()
            ordine = request.form.get('ordine', type=int)
            etichetta = request.form.get('etichetta', '').strip()
            
            if not all((testo, area, ordine)):
                flash('Tutti i campi sono obbligatori', 'error')
                return redirect(url_for('tabella_risposte'))

            # Verifica se esiste già una domanda con lo stesso ordine
            if Domanda.query.filter_by(ordine=ordine).first():
                flash('Esiste già una domanda con questo numero d\'ordine', 'error')
                return redirect(url_for('tabella_risposte'))
            
            domanda = Domanda(
                testo=testo,
                area=area,
                ordine=ordine,
                etichetta=etichetta
            )
            db.session.add(domanda)
            db.session.commit()

            # Aggiorna la MAPPA_DOMANDE
            MAPPA_DOMANDE[ordine] = testo

            # Aggiorna AREE e AREE_INVERSE
            if ordine not in AREE_INVERSE:
                if area in AREE:
                    AREE[area].append(ordine)
                else:
                    AREE[area] = [ordine]
                AREE_INVERSE[ordine] = area

            flash('Domanda aggiunta con successo', 'success')
            
        elif azione == 'modifica_domanda':
            domanda_id = request.form.get('domanda_id', type=int)
            domanda = Domanda.query.get_or_404(domanda_id)
            
            testo = request.form.get('testo', '').strip()
            area = request.form.get('area', '').strip()
            ordine = request.form.get('ordine', type=int)
            etichetta = request.form.get('etichetta', '').strip()
            
            if not all((testo, area, ordine)):
                flash('Tutti i campi sono obbligatori', 'error')
                return redirect(url_for('tabella_risposte'))

            # Se l'ordine è cambiato, verifica che non esista già
            if ordine != domanda.ordine and Domanda.query.filter_by(ordine=ordine).first():
                flash('Esiste già una domanda con questo numero d\'ordine', 'error')
                return redirect(url_for('tabella_risposte'))

            # Aggiorna MAPPA_DOMANDE, AREE e AREE_INVERSE se necessario
            if ordine != domanda.ordine:
                # Rimuovi vecchio ordine
                if domanda.ordine in MAPPA_DOMANDE:
                    del MAPPA_DOMANDE[domanda.ordine]
                if domanda.ordine in AREE_INVERSE:
                    AREE[domanda.area].remove(domanda.ordine)
                    del AREE_INVERSE[domanda.ordine]
                
                # Aggiungi nuovo ordine
                MAPPA_DOMANDE[ordine] = testo
                if area in AREE:
                    AREE[area].append(ordine)
                else:
                    AREE[area] = [ordine]
                AREE_INVERSE[ordine] = area
            else:
                # Aggiorna solo il testo se l'ordine non è cambiato
                MAPPA_DOMANDE[ordine] = testo
            
            domanda.testo = testo
            domanda.area = area
            domanda.ordine = ordine
            domanda.etichetta = etichetta
            
            db.session.commit()
            flash('Domanda modificata con successo', 'success')
            
        elif azione == 'elimina_domanda':
            domanda_id = request.form.get('domanda_id', type=int)
            domanda = Domanda.query.get_or_404(domanda_id)

            # Rimuovi da MAPPA_DOMANDE, AREE e AREE_INVERSE
            if domanda.ordine in MAPPA_DOMANDE:
                del MAPPA_DOMANDE[domanda.ordine]
            if domanda.ordine in AREE_INVERSE:
                AREE[domanda.area].remove(domanda.ordine)
                del AREE_INVERSE[domanda.ordine]
            
            db.session.delete(domanda)
            db.session.commit()
            flash('Domanda eliminata con successo', 'success')
        
        return redirect(url_for('tabella_risposte'))

    # Get all responses for GET request
    risposte = Risposta.query.order_by(Risposta.timestamp.desc()).all()
    
    # Get all questions for question management
    domande = Domanda.query.order_by(Domanda.ordine).all()
    
    # Prepare headers and data for the table
    headers = ['ID', 'Data']
    headers.extend([f'Q{i}' for i in range(1, 61)])
    headers.extend(['Media Motivazione', 'Media Risorse', 'Media Elaborazione', 
                   'Media Tempo', 'Media Strategie', 'Media Concentrazione',
                   'Media Selezione', 'Media Atteggiamento', 'Media Monitoraggio',
                   'Media Ansia'])
    
    data = []
    for r in risposte:
        row = [r.id, r.timestamp.strftime('%d/%m/%Y %H:%M')]
        row.extend([r.risposte.get(f'q{i}', '') for i in range(1, 61)])
        row.extend([
            round(r.media_motivazione, 2),
            round(r.media_risorse, 2),
            round(r.media_elaborazione, 2),
            round(r.media_tempo, 2),
            round(r.media_strategie, 2),
            round(r.media_concentrazione, 2),
            round(r.media_selezione, 2),
            round(r.media_atteggiamento, 2),
            round(r.media_monitoraggio, 2),
            round(r.media_ansia, 2)
        ])
        data.append(row)
    
    return render_template('admin/tabella_risposte.html',
                         headers=headers,
                         data=data,
                         risposte=risposte,
                         mappa_domande=MAPPA_DOMANDE,
                         aree_domande={str(k): v for k, v in AREE_INVERSE.items()},
                         domande=domande,
                         aree=AREE.keys())

@app.route('/admin/download_risultati/<int:risposta_id>')
@login_required
def download_risultati(risposta_id):
    """Genera un file PDF con i risultati di una specifica risposta"""
    # Recupera la risposta dal database
    risposta = Risposta.query.get_or_404(risposta_id)
    
    # Crea un buffer di memoria per il PDF
    buffer = io.BytesIO()
    
    # Crea il documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    # Intestazione del documento
    title_style = styles['Title']
    title = Paragraph("Risultato Questionario di Autovalutazione", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Data compilazione
    date_style = styles['Normal']
    date_style.alignment = 1  # Centrato
    date_text = Paragraph(f"Data compilazione: {risposta.timestamp.strftime('%d/%m/%Y %H:%M')}", date_style)
    elements.append(date_text)
    elements.append(Spacer(1, 20))
    
    # Tabella con i punteggi per area
    heading_style = styles['Heading2']
    elements.append(Paragraph("Punteggi per Area", heading_style))
    elements.append(Spacer(1, 10))
    
    data = [
        ["Area", "Punteggio"],
        ["Motivazione", f"{risposta.media_motivazione:.2f}"],
        ["Uso delle Risorse Accademiche", f"{risposta.media_risorse:.2f}"],
        ["Elaborazione delle Informazioni", f"{risposta.media_elaborazione:.2f}"],
        ["Gestione del Tempo", f"{risposta.media_tempo:.2f}"],
        ["Strategie per Svolgere Prove", f"{risposta.media_strategie:.2f}"],
        ["Concentrazione", f"{risposta.media_concentrazione:.2f}"],
        ["Selezione dei Concetti Principali", f"{risposta.media_selezione:.2f}"],
        ["Atteggiamento", f"{risposta.media_atteggiamento:.2f}"],
        ["Auto-Monitoraggio", f"{risposta.media_monitoraggio:.2f}"],
        ["Ansia", f"{risposta.media_ansia:.2f}"]
    ]
    
    table = Table(data, colWidths=[300, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Feedback per area
    elements.append(Paragraph("Feedback per Area", heading_style))
    elements.append(Spacer(1, 10))
    
    # Creo uno stile specifico per i feedback con allineamento a sinistra
    feedback_style = ParagraphStyle(
        'FeedbackStyle',
        parent=styles['Normal'],
        alignment=0  # 0 = sinistra in ReportLab
    )
    
    # Recupera i feedback per ciascuna area
    feedback_motivazione = get_feedback(risposta.media_motivazione, 'MOTIVAZIONE')
    feedback_risorse = get_feedback(risposta.media_risorse, 'USO DELLE RISORSE ACCADEMICHE')
    feedback_elaborazione = get_feedback(risposta.media_elaborazione, 'ELABORAZIONE DELLE INFORMAZIONI')
    feedback_tempo = get_feedback(risposta.media_tempo, 'GESTIONE DEL TEMPO')
    feedback_strategie = get_feedback(risposta.media_strategie, 'STRATEGIE PER SVOLGERE PROVE')
    feedback_concentrazione = get_feedback(risposta.media_concentrazione, 'CONCENTRAZIONE')
    feedback_selezione = get_feedback(risposta.media_selezione, 'SELEZIONE DEI CONCETTI PRINCIPALI')
    feedback_atteggiamento = get_feedback(risposta.media_atteggiamento, 'ATTEGGIAMENTO')
    feedback_monitoraggio = get_feedback(risposta.media_monitoraggio, 'AUTO-MONITORAGGIO')
    feedback_ansia = get_feedback(risposta.media_ansia, 'ANSIA')
    
    # Aggiunge tutti i feedback al PDF
    for area, feedback in [
        ("Motivazione", feedback_motivazione),
        ("Uso delle Risorse Accademiche", feedback_risorse),
        ("Elaborazione delle Informazioni", feedback_elaborazione),
        ("Gestione del Tempo", feedback_tempo),
        ("Strategie per Svolgere Prove", feedback_strategie),
        ("Concentrazione", feedback_concentrazione),
        ("Selezione dei Concetti Principali", feedback_selezione),
        ("Atteggiamento", feedback_atteggiamento),
        ("Auto-Monitoraggio", feedback_monitoraggio),
        ("Ansia", feedback_ansia)
    ]:
        area_style = styles['Heading3']
        elements.append(Paragraph(area, area_style))
        elements.append(Paragraph(feedback, feedback_style))  # Uso il nuovo stile per i feedback
        elements.append(Spacer(1, 10))
    
    # Genera il PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Invia il file
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'risultati_questionario_{risposta_id}.pdf'
    )

def get_stats(risposte):
    if not risposte:
        return {
            'media_motivazione': 0,
            'media_risorse': 0,
            'media_elaborazione': 0,
            'media_tempo': 0,
            'media_strategie': 0,
            'media_concentrazione': 0,
            'media_selezione': 0,
            'media_atteggiamento': 0,
            'media_monitoraggio': 0,
            'media_ansia': 0,
            'totale_risposte': 0
        }
    
    return {
        'media_motivazione': sum(r.media_motivazione for r in risposte) / len(risposte),
        'media_risorse': sum(r.media_risorse for r in risposte) / len(risposte),
        'media_elaborazione': sum(r.media_elaborazione for r in risposte) / len(risposte),
        'media_tempo': sum(r.media_tempo for r in risposte) / len(risposte),
        'media_strategie': sum(r.media_strategie for r in risposte) / len(risposte),
        'media_concentrazione': sum(r.media_concentrazione for r in risposte) / len(risposte),
        'media_selezione': sum(r.media_selezione for r in risposte) / len(risposte),
        'media_atteggiamento': sum(r.media_atteggiamento for r in risposte) / len(risposte),
        'media_monitoraggio': sum(r.media_monitoraggio for r in risposte) / len(risposte),
        'media_ansia': sum(r.media_ansia for r in risposte) / len(risposte),
        'totale_risposte': len(risposte)
    }

def get_area_domanda(numero):
    """Helper per ottenere l'area di una domanda"""
    for area, numeri in AREE.items():
        if numero in numeri:
            return area
    return None

def export_to_excel():
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Risposte Questionario"
    
    # Ottengo i dati
    headers = ['ID', 'Data']
    headers.extend([f'Q{i}' for i in range(1, 61)])
    headers.extend(['Motivazione', 'Risorse', 'Elaborazione', 'Tempo', 
                   'Strategie', 'Concentrazione', 'Selezione', 
                   'Atteggiamento', 'Monitoraggio', 'Ansia'])
    
    # Formattazione intestazioni
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Recupero e formatto i dati
    risposte = Risposta.query.order_by(Risposta.timestamp.desc()).all()
    for row_num, risposta in enumerate(risposte, 2):
        row = [
            risposta.id,
            risposta.timestamp.strftime('%d/%m/%Y %H:%M')
        ]
        
        # Aggiungo le risposte individuali
        for i in range(1, 61):
            row.append(risposta.risposte.get(f'q{i}', ''))
        
        # Aggiungo le medie
        row.extend([
            round(risposta.media_motivazione, 2),
            round(risposta.media_risorse, 2),
            round(risposta.media_elaborazione, 2),
            round(risposta.media_tempo, 2),
            round(risposta.media_strategie, 2),
            round(risposta.media_concentrazione, 2),
            round(risposta.media_selezione, 2),
            round(risposta.media_atteggiamento, 2),
            round(risposta.media_monitoraggio, 2),
            round(risposta.media_ansia, 2)
        ])
        
        # Inserisco i dati nel foglio
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            if col_num > 62:  # Celle delle medie
                cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center')
    
    # Secondo foglio con l'elenco delle domande
    worksheet2 = workbook.create_sheet(title="Elenco Domande")
    worksheet2.append(['Numero', 'Area', 'Testo Domanda'])
    
    for numero, testo in MAPPA_DOMANDE.items():
        area = get_area_domanda(numero)
        worksheet2.append([numero, area, testo])
    
    workbook.save(output)
    output.seek(0)
    return output

def init_default_clusters():
    """Inizializza i cluster predefiniti per ogni area con feedback specifici."""
    default_clusters = {
        'MOTIVAZIONE': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'La tua motivazione allo studio appare bassa. Suggerimenti: 1) Stabilisci obiettivi di studio a breve termine più raggiungibili; 2) Celebra i piccoli successi quotidiani; 3) Ricorda i tuoi obiettivi di lungo termine e perché hai scelto questo percorso di studi.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'La tua motivazione è nella media. Stai facendo un buon lavoro, ma potresti migliorare: 1) Definisci obiettivi più specifici e misurabili; 2) Crea connessioni tra lo studio e i tuoi interessi personali; 3) Mantieni un diario dei progressi.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Hai un\'ottima motivazione allo studio! Continua così: 1) Mantieni questo atteggiamento positivo; 2) Condividi la tua esperienza con i compagni di studio; 3) Sfidati con obiettivi sempre più ambiziosi mantenendo l\'equilibrio.'
            }
        ],
        'USO DELLE RISORSE ACCADEMICHE': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'Utilizzi poco le risorse accademiche disponibili. Suggerimenti: 1) Esplora i servizi di supporto offerti dall\'università; 2) Partecipa alle ore di ricevimento dei docenti; 3) Unisciti a gruppi di studio.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'Fai un uso moderato delle risorse accademiche. Per migliorare: 1) Incrementa la frequenza di utilizzo del tutorato; 2) Sfrutta maggiormente la biblioteca; 3) Partecipa più attivamente alle attività di laboratorio.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Ottimo utilizzo delle risorse accademiche! Continua a: 1) Mantenere questo approccio proattivo; 2) Condividere le tue esperienze positive con altri studenti; 3) Esplorare risorse aggiuntive per approfondire i tuoi interessi.'
            }
        ],
        'ELABORAZIONE DELLE INFORMAZIONI': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'La tua elaborazione delle informazioni necessita miglioramenti. Prova a: 1) Creare mappe concettuali; 2) Riformulare i concetti con parole tue; 3) Cercare esempi pratici per i concetti teorici.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'La tua capacità di elaborazione è nella media. Per migliorare: 1) Approfondisci le connessioni tra diversi argomenti; 2) Pratica più spesso la spiegazione dei concetti ad altri; 3) Crea riassunti più strutturati.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Eccellente capacità di elaborazione! Mantieni questo livello: 1) Continua a creare collegamenti interdisciplinari; 2) Sviluppa metodi personali di sintesi; 3) Aiuta altri studenti a migliorare le loro strategie.'
            }
        ],
        'GESTIONE DEL TEMPO': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'La tua gestione del tempo richiede attenzione. Suggerimenti: 1) Crea un calendario di studio settimanale; 2) Usa timer per le sessioni di studio; 3) Elimina le principali fonti di distrazione.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'Gestisci il tempo in modo discreto. Per migliorare: 1) Perfeziona la pianificazione settimanale; 2) Bilancia meglio studio e pause; 3) Identifica e sfrutta le tue ore di massima produttività.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Ottima gestione del tempo! Continua a: 1) Mantenere questa efficiente organizzazione; 2) Adattare il piano di studio quando necessario; 3) Aiutare altri a migliorare la loro gestione del tempo.'
            }
        ],
        'STRATEGIE PER SVOLGERE PROVE': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'Le tue strategie per le prove necessitano miglioramenti. Prova a: 1) Esercitarti con prove degli anni precedenti; 2) Simulare le condizioni d\'esame; 3) Analizzare gli errori commessi nelle prove precedenti.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'Le tue strategie per le prove sono nella media. Per migliorare: 1) Diversifica le modalità di esercitazione; 2) Organizza meglio il tempo durante le prove; 3) Sviluppa strategie specifiche per diversi tipi di esame.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Eccellenti strategie per le prove! Continua a: 1) Perfezionare il tuo approccio; 2) Adattare le strategie ai diversi tipi di esame; 3) Condividere i tuoi metodi efficaci con altri studenti.'
            }
        ],
        'CONCENTRAZIONE': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'La tua concentrazione richiede miglioramenti. Suggerimenti: 1) Trova un ambiente di studio privo di distrazioni; 2) Usa la tecnica del pomodoro (25 minuti di studio, 5 di pausa); 3) Pratica esercizi di mindfulness.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'La tua concentrazione è nella media. Per migliorare: 1) Aumenta gradualmente i tempi di studio concentrato; 2) Identifica i momenti della giornata più produttivi; 3) Mantieni un ambiente di studio più organizzato.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Ottima capacità di concentrazione! Continua a: 1) Mantenere queste buone abitudini; 2) Alternare periodi di concentrazione intensa e pause; 3) Variare le attività per mantenere alta l\'attenzione.'
            }
        ],
        'SELEZIONE DEI CONCETTI PRINCIPALI': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'La tua capacità di selezione dei concetti chiave necessita attenzione. Prova a: 1) Utilizzare l\'indice del libro come guida; 2) Evidenziare le parole chiave durante la lettura; 3) Creare schemi gerarchici dei contenuti.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'La tua selezione dei concetti principali è nella media. Per migliorare: 1) Pratica più sintesi dei contenuti; 2) Confronta gli appunti con i colleghi; 3) Crea mappe mentali più dettagliate.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Eccellente capacità di selezione dei concetti! Continua a: 1) Affinare le tue tecniche di sintesi; 2) Creare collegamenti tra concetti chiave; 3) Aiutare altri a sviluppare questa competenza.'
            }
        ],
        'ATTEGGIAMENTO': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'Il tuo atteggiamento verso lo studio potrebbe migliorare. Suggerimenti: 1) Identifica gli aspetti positivi del tuo percorso; 2) Cerca ispirazione da studenti motivati; 3) Concentrati sui progressi più che sulle difficoltà.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'Il tuo atteggiamento è nella media. Per migliorare: 1) Sviluppa una mentalità più orientata alla crescita; 2) Valorizza maggiormente i tuoi successi; 3) Vedi gli errori come opportunità di apprendimento.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Ottimo atteggiamento verso lo studio! Continua a: 1) Mantenere questo approccio positivo; 2) Ispirare gli altri con il tuo esempio; 3) Cercare nuove sfide per la tua crescita.'
            }
        ],
        'AUTO-MONITORAGGIO': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'Le tue capacità di auto-monitoraggio necessitano attenzione. Prova a: 1) Tenere un diario di studio; 2) Verificare regolarmente la comprensione; 3) Stabilire momenti specifici per l\'autovalutazione.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'Le tue capacità di auto-monitoraggio sono nella media. Per migliorare: 1) Aumenta la frequenza delle autovalutazioni; 2) Usa strumenti più strutturati per il monitoraggio; 3) Confronta le tue percezioni con i risultati effettivi.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Eccellenti capacità di auto-monitoraggio! Continua a: 1) Perfezionare i tuoi metodi di autovalutazione; 2) Adattare le strategie in base ai risultati; 3) Condividere le tue tecniche efficaci.'
            }
        ],
        'ANSIA': [
            {
                'min_valore': 1.0,
                'max_valore': 2.5,
                'feedback': 'Il tuo livello di ansia sembra gestibile. Continua a: 1) Mantenere un approccio calmo allo studio; 2) Utilizzare tecniche di rilassamento quando necessario; 3) Preparati in modo adeguato per aumentare la sicurezza.'
            },
            {
                'min_valore': 2.5,
                'max_valore': 3.5,
                'feedback': 'Il tuo livello di ansia è moderato. Suggerimenti: 1) Pratica tecniche di respirazione; 2) Organizza meglio il tempo di studio; 3) Concentrati sulla preparazione invece che sulle preoccupazioni.'
            },
            {
                'min_valore': 3.5,
                'max_valore': 5.0,
                'feedback': 'Il tuo livello di ansia richiede attenzione. Prova a: 1) Parlare con il servizio di counseling universitario; 2) Imparare tecniche di gestione dello stress; 3) Sviluppare una routine di preparazione rassicurante.'
            }
        ]
    }

    # Elimina i cluster esistenti
    ClusterRisposte.query.delete()
    
    # Inserisci i nuovi cluster predefiniti
    for area, clusters in default_clusters.items():
        for cluster in clusters:
            new_cluster = ClusterRisposte(
                area=area,
                min_valore=cluster['min_valore'],
                max_valore=cluster['max_valore'],
                feedback=cluster['feedback']
            )
            db.session.add(new_cluster)
    
    db.session.commit()

def init_db():
    with app.app_context():
        db.create_all()
        
        # Crea utente admin se non esiste
        if not User.query.filter_by(username='admin').first():
            admin_user = User(
                username='admin',
                password_hash=generate_password_hash('admin123')
            )
            db.session.add(admin_user)
            db.session.commit()
        
        # Popola le domande se non esistono
        if Domanda.query.count() == 0:
            for numero, testo in MAPPA_DOMANDE.items():
                area = AREE_INVERSE[numero]
                domanda = Domanda(
                    testo=testo,
                    area=area,
                    ordine=numero
                )
                db.session.add(domanda)
            db.session.commit()
        
        # Inizializza i cluster predefiniti se non esistono
        if ClusterRisposte.query.count() == 0:
            init_default_clusters()

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5005, host='0.0.0.0')